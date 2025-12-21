from django.core.management.base import BaseCommand
from api.models import StockInfo
from pathlib import Path


class Command(BaseCommand):
    help = 'Load stock short codes and names from .xlsx files in a directory'

    def add_arguments(self, parser):
        parser.add_argument('--dir', type=str, default='stock_info', help='Directory containing .xlsx files')
        parser.add_argument('--head', type=int, default=0, help='If >0, print first N parsed rows and exit (no DB writes)')

    def handle(self, *args, **options):
        base = Path(options['dir'])
        created = 0
        updated = 0
        head_n = int(options.get('head', 0) or 0)
        printed = 0

        # collect all .xlsx files in directory
        files = sorted(base.glob('*.xlsx'))
        if not files:
            self.stdout.write(self.style.WARNING(f'No .xlsx files found in {base}'))
            return

        try:
            from openpyxl import load_workbook
        except Exception:
            self.stdout.write(self.style.ERROR('openpyxl is required. Install with `pip install openpyxl`'))
            return

        for f in files:
            fname = f.name.lower()
            if 'kospi' in fname:
                market = 'KOSPI'
            elif 'kosdaq' in fname:
                market = 'KOSDAQ'
            else:
                market = 'UNKNOWN'


            wb = load_workbook(filename=str(f), read_only=True, data_only=True)
            # iterate first sheet only
            sheet = wb[wb.sheetnames[0]]

            # rely on explicit headers in the Excel files; remove regex heuristics

            iterator = sheet.iter_rows(values_only=True)
            try:
                header_row = next(iterator)
            except StopIteration:
                continue

            # normalize header cells
            headers = [str(h).strip() if h is not None else '' for h in header_row]
            lower_headers = [h.lower() for h in headers]

            # candidate header names
            short_header_candidates = ['단축코드', '단축 코드', 'short_code', 'shortcode', '단축']
            name_header_candidates = ['한글종목명', '한글 종목명', '한글명', '종목명', 'name', '한글']

            def find_index(candidates):
                for cand in candidates:
                    # exact match (Korean)
                    if cand in headers:
                        return headers.index(cand)
                # try lowercase variants
                for cand in candidates:
                    if cand.lower() in lower_headers:
                        return lower_headers.index(cand.lower())
                return None

            idx_short = find_index(short_header_candidates)
            idx_name = find_index(name_header_candidates)

            # if headers not found, we'll fallback to heuristics per-row
            for row in iterator:
                if not row:
                    continue
                # normalize cells to strings
                cells = [str(c).strip() if c is not None else '' for c in row]
                if not any(cells):
                    continue

                # prefer header-based extraction (Excel files contain proper columns)
                short_code = ''
                name = ''

                if idx_short is not None and idx_short < len(cells):
                    short_code = cells[idx_short]
                if idx_name is not None and idx_name < len(cells):
                    name = cells[idx_name]

                # simple fallbacks when headers are missing: use first columns
                if not short_code:
                    short_code = cells[0] if len(cells) > 0 and cells[0] else ''

                if not name:
                    # prefer the next non-empty cell that isn't the short_code
                    name = next((c for c in cells[1:] if c and c != short_code), '')

                if not short_code or not name:
                    continue

                # truncate to model limits
                short_code = short_code[:64]
                name = name[:255]

                if head_n > 0:
                    self.stdout.write(f"{f.name}\t{market}\tshort_code={short_code}\tname={name}")
                    printed += 1
                    if printed >= head_n:
                        self.stdout.write(self.style.SUCCESS(f"Preview printed {printed} rows."))
                        return
                    continue

                obj, created_flag = StockInfo.objects.update_or_create(
                    short_code=short_code,
                    defaults={
                        'name': name,
                        'market': market,
                    }
                )
                if created_flag:
                    created += 1
                else:
                    updated += 1

        self.stdout.write(self.style.SUCCESS(f"Import complete: {created} created, {updated} updated"))
